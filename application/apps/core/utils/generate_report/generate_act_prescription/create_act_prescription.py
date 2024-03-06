import openpyxl
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from apps.MyBot import bot_send_message
from apps.core.bot.messages.messages import Messages
from apps.core.database.db_utils import db_get_clean_headers
from apps.core.utils.generate_report.generate_act_prescription.set_act_format_ import (format_act_photo_description,
                                                                                       format_act_photo_header)
from apps.core.utils.generate_report.generate_act_prescription.set_act_values import (set_act_photographic_materials)
from apps.core.utils.secondary_functions.get_filepath import get_image_name
from config.config import Udocan_media_path
from loader import logger


async def create_xlsx(chat_id: int, full_act_path: str):
    """
    """

    is_created: bool = await create_new_xlsx(report_file=full_act_path)
    if is_created is None:
        logger.warning(Messages.Error.workbook_not_create)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_create)
        return

    workbook = await get_workbook(fill_report_path=full_act_path)
    if workbook is None:
        logger.warning(Messages.Error.workbook_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_found)
        return

    worksheet = await get_worksheet(workbook, index=0)
    if worksheet is None:
        logger.warning(Messages.Error.worksheet_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.worksheet_not_found)
        return

    return workbook, worksheet


async def create_new_xlsx(report_file: str) -> bool:
    """Создание xlsx
    """
    try:
        wb = openpyxl.Workbook()
        wb.save(report_file)
        return True

    except Exception as err:
        logger.error(F"set_border {repr(err)}")
        return False


async def get_worksheet(wb: Workbook, index: int = 0) -> Worksheet:
    """Получение Страницы из документа по индексу
    :param wb: Workbook - книга xls
    :param index: int - индекс листа
    :return: worksheet or None
    """
    try:
        worksheet: Worksheet = wb.worksheets[index]
        return worksheet

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def get_workbook(fill_report_path: str) -> Workbook:
    """Открыть и загрузить Workbook
    :param fill_report_path: str полный путь к файлу
    :return: Workbook or None
    """
    try:
        workbook: Workbook = openpyxl.load_workbook(fill_report_path)
        return workbook

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def anchor_photo(dataframe, row_number, workbook, worksheet, full_act_path, act_date=None):
    """Вставка фото в ячейку с учетом смещения. Возвращает область печати с учетом фото

    """
    photo_row: int = 63 + row_number
    row_num: int = 0
    img_params: dict = {"height": 220,
                        "anchor": True,
                        "scale": True, }

    row_act_header: int = photo_row - 1
    row_act_photo: int = photo_row
    img_params["row_header"] = row_act_header
    img_params["row_description"] = row_act_photo
    img_params["row"] = row_act_photo

    table_name: str = 'core_violations'
    clean_headers: list = await db_get_clean_headers(table_name=table_name)

    for num_data, violation_data in enumerate(dataframe.itertuples(index=False), start=1):
        # for num_data, violation_data in enumerate(df.to_dict('index'), start=1):

        violation_dict = dict(zip(clean_headers, violation_data))

        # img_params["photo_full_name"] = str (Udocan_media_path) + "HSE" + os.sep + str(violation_dict.get('photo', None))
        # img_params["photo_full_name"] = await get_photo_full_name(
        #     media_path=await get_directory_name(Udocan_media_path, "HSE"),
        #     photo_path=await get_image_name(violation_dict.get('photo', None))
        # )
        img_params["photo_full_name"] = violation_dict.get('photo_full_name', None)
        if num_data % 2 != 0:
            img_params["column"] = 'B'
            img_params["column_img"] = 2
            img_params["column_description"] = 6
            img_params["description"] = violation_dict.get('description', None)

            if num_data != 1:
                img_params["row_header"] = img_params["row_header"] + 2
                img_params["row_description"] = img_params["row_description"] + 2
                img_params["row"] = img_params["row"] + 2
                await format_act_photo_header(worksheet, img_params["row_header"])
                await format_act_photo_description(worksheet, img_params["row_description"])
                workbook.save(full_act_path)
                photo_row += 2
        else:
            img_params["column"] = 'H'
            img_params["column_img"] = 8
            img_params["column_description"] = 12
            img_params["description"] = violation_dict.get('description', None)

        is_anchor: bool = False
        try:
            is_anchor = await set_act_photographic_materials(worksheet, img_params)

        except Exception as err:
            logger.error(f"set_act_photographic_materials {num_data= } {row_num=} {repr(err)} ")

        if is_anchor:
            workbook.save(full_act_path)
            # set header
            worksheet.cell(row=img_params["row_header"],
                           column=img_params["column_img"],
                           value=f'Фото {num_data}')
            workbook.save(full_act_path)

            # set description
            worksheet.cell(row=img_params["row_description"],
                           column=img_params["column_description"],
                           value=img_params["description"])
            workbook.save(full_act_path)

    print_area = f'$A$1:M{photo_row + row_num + 1}'

    return print_area


async def insert_service_image(worksheet: Worksheet, *, chat_id: int = None, service_image_name: str = 'Logo',
                               img_params: dict = None) -> bool:
    """Вставка изображений в файл

    :param service_image_name: str - имя файла для обработки
    :param chat_id: int - id пользователя (папки) где находится logo
    :param img_params: dict параметры вставки
    :param worksheet:
    :return: bool
    """

    photo_full_name: str = await get_image_name(Udocan_media_path, "HSE", str(chat_id), f"{service_image_name}.jpg")

    # if chat_id:
    #     photo_full_name: str = await get_image_name(Udocan_media_path, "HSE",
    #     str(chat_id), f"{service_image_name}.jpg")

    if not os.path.isfile(photo_full_name):
        logger.error("service image not found")
        photo_full_name: str = await get_image_name(Udocan_media_path, "HSE", str(chat_id), "Logo.jpg")

    if not img_params:
        img_params: dict = {
            'photo_full_name': photo_full_name,
            "height": 90,
            "width": 230,
            "anchor": True,
            "column": 'B',
            "column_img": 2,
            "row": 2,
        }

    if not os.path.isfile(photo_full_name):
        logger.error("service image not found")
        return False

    img: Image = Image(photo_full_name)

    img = await image_preparation(img, img_params)

    result = await insert_images(worksheet, img=img)
    if result:
        return True
    return False


async def image_preparation(img: Image, img_params: dict):
    """Подготовка изображения перед вставкой на страницу
    """

    height = img_params.get("height")
    scale = img_params.get("scale")
    width = img_params.get("width")
    anchor = img_params.get("anchor")
    column = img_params.get("column")
    row = img_params.get("row")

    # высота изображения
    if height:
        img.height = height

    # ширина изображения
    if width:
        img.width = width

    # изменение пропорций изображения
    if scale:
        scale = img.height / max(img.height, img.width)
        img.width = img.width * scale

    # прикрепление изображение по адресу str(column) + str(row)
    if anchor:
        img.anchor = str(column) + str(row)

    return img


async def insert_images(worksheet: Worksheet, img: Image) -> bool:
    """Вставка изображения на лист worksheet*,
    :param img: файл изображения
    :param worksheet: Worksheet - страница документа для вставки изображения
    :return:
    """

    try:
        worksheet.add_image(img)
        return True

    except Exception as err:
        logger.error(f"insert_images {repr(err)}")
        return False


async def set_act_body_values(worksheet):
    """

    :param worksheet:
    :return:
    """
    values = [
        {'coordinate': 'B6', 'value': 'Акт предписание № от', 'row': '6', 'column': '2'},
        {'coordinate': 'B7', 'value': 'об устранении нарушений ', 'row': '7', 'column': '2'},

        {'coordinate': 'B9', 'value': 'тут наименование организации', 'row': '9', 'column': '2'},
        {'coordinate': 'B10',
         'value': '(указать кому адресовано (полное или сокращенное наименование юридического '
                  'лица либо индивидуального предпринимателя, ИНН) ',
         'row': '10', 'column': '2'},
        {'coordinate': 'B11', 'value': 'Мною:', 'row': '11', 'column': '2'},
        {'coordinate': 'B12', 'value': 'тут должность и ФИО полностью выдавшего', 'row': '12', 'column': '2'},

        {'coordinate': 'B14',
         'value': '(фамилия, имя, отчество (последнее – при наличии), должность должностного '
                  'лица, уполномоченного выдавать предписания',
         'row': '14', 'column': '2'},
        {'coordinate': 'B15', 'value': 'проведена проверка по соблюдению требований ОТ, ПБ и ООС, в отношении:',
         'row': '15', 'column': '2'},
        {'coordinate': 'B16', 'value': 'тут наименование организации', 'row': '16', 'column': '2'},
        {'coordinate': 'B17',
         'value': '(указать полное наименование юридического лица либо индивидуального предпринимателя)', 'row': '17',
         'column': '2'},

        {'coordinate': 'B19', 'value': 'В присутствии:', 'row': '19', 'column': '2'},
        {'coordinate': 'B20', 'value': 'тут ответственный', 'row': '20', 'column': '2'},

        {'coordinate': 'B22', 'value': 'ПРЕДПИСЫВАЕТСЯ ', 'row': '22', 'column': '2'},
        {'coordinate': 'B23', 'value': 'Принять меры по устранению выявленных нарушений в установленные сроки.',
         'row': '23', 'column': '2'},

        {'coordinate': 'B25', 'value': '№', 'row': '25', 'column': '2'},
        {'coordinate': 'B25', 'value': 'Описание и характер выявленных нарушений', 'row': '25', 'column': '3'},
        {'coordinate': 'B25',
         'value': 'Наименование НПА, номера подпунктов, пунктов, требования которых нарушены или не соблюдены',
         'row': '25', 'column': '6'},
        {'coordinate': 'B25', 'value': 'Предписываемые меры по устранению выявленного нарушения', 'row': '25',
         'column': '9'}, {'coordinate': 'B25', 'value': 'Срок устранения нарушений', 'row': '25', 'column': '12'},
        {'coordinate': 'B26', 'value': 'п/п', 'row': '26', 'column': '2'},
        {'coordinate': 'B27', 'value': 'тут наименование ПО', 'row': '27', 'column': '2'},
        {'coordinate': 'B28', 'value': '1', 'row': '28', 'column': '2'},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return None


async def set_act_page_after_footer_setup(worksheet: Worksheet, print_area: str, break_line: int = None) -> bool:
    """Установка параметров страницы

    :param break_line: int
    :param print_area: str
    :param worksheet: Worksheet
    :return: bool
    """

    #  https://xlsxons.horizontalCentered = True
    worksheet.print_area = print_area

    #  масштабный коэффициент для распечатываемой страницы
    # worksheet.set_print_scale(75)

    if break_line:
        worksheet.row_breaks.append(Break(id=break_line))
        # worksheet.col_breaks.append(Break(id=13))

    return True
