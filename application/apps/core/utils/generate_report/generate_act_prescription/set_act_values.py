import os
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.worksheet import Worksheet

from pandas import DataFrame

from loader import logger
from apps.core.database.db_utils import db_get_data_dict_from_table_with_id


async def set_act_violation_values(worksheet: Worksheet, dataframe: DataFrame, workbook: Workbook,
                                   full_act_path: str) -> int:
    """Заполнение акта значениями из dataframe. Сортировка по main_location_id и sub_locations_id

    :param full_act_path:
    :param workbook:
    :param worksheet: страница для заполнения
    :param dataframe: dataframe с данными нарушений
    :return: None
    """

    # берём уникальные значения main_location_id
    unique_main_locations_ids: list = dataframe.main_location_id.unique().tolist()
    row_number: int = 0
    item_row_value: int = 0

    for main_location_id in unique_main_locations_ids:
        new_header_row = True

        # разделяем dataframe по каждому уникальному значению
        main_location_values: dataframe = dataframe.loc[dataframe['main_location_id'] == main_location_id]
        if main_location_values.empty:
            continue

        # разделяем по sub_location
        unique_sub_locations_ids: list = main_location_values.sub_location_id.unique().tolist()

        for sub_locations_id in unique_sub_locations_ids:

            new_header_row = True
            sub_locations_values: dataframe = main_location_values.loc[dataframe['sub_location_id'] == sub_locations_id]
            if sub_locations_values.empty:
                continue

            # итерируемся по dataframe как по tuple
            for row in sub_locations_values.itertuples(index=False):
                item_row_value += 1
                row_number = await set_act_worksheet_cell_value(
                    worksheet, row, row_number, new_header_row, workbook, full_act_path, item_row_value
                )
                new_header_row = False
                row_number += 1

    return row_number


async def set_act_worksheet_cell_value(worksheet: Worksheet, violation_values: DataFrame, row_number: int,
                                       new_header_row: bool, workbook: Workbook, full_act_path: str,
                                       item_row_value: int) -> int:
    """Заполнение тела акта каждым пунктом

        :param item_row_value:
        :param full_act_path:
        :param workbook:
        :param new_header_row: bool  Требуется ли новый заголовок?
        :param row_number:
        :param violation_values: DataFrame
        :param worksheet: страница для заполнения
        :return
        """

    if row_number == 0:
        logger.debug(violation_values)

        await set_single_violation(worksheet, violation_values)
        workbook.save(full_act_path)
        return row_number

    if new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_value_title(worksheet, violation_values, row_number)
        workbook.save(full_act_path)
        row_number += 1

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number

    if not new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number


async def set_act_violation(worksheet: Worksheet, violation_values: DataFrame, row_number: int,
                            item_row_value: int) -> int:
    """

    :param item_row_value:
    :param row_number:
    :param worksheet:
    :param violation_values:
    :return:
    """

    row_value = 28 + row_number

    worksheet.cell(row=row_value, column=2, value=item_row_value)
    normative_document: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=row_value, column=3, value=violation_values.description)
        # worksheet.cell(row=row_value, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=row_value, column=3, value=normative_document['title'])
        # worksheet.cell(row=row_value, column=15, value=normative_document['title'])

    worksheet.cell(row=row_value, column=6, value=normative_document['normative'])
    # worksheet.cell(row=row_value, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in [None, '', ' ', '.', '*', '/']:
        worksheet.cell(row=row_value, column=9, value=violation_values.comment)
        # worksheet.cell(row=row_value, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=row_value, column=9, value=normative_document['procedure'])
        # worksheet.cell(row=row_value, column=17, value=normative_document['procedure'])

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=row_value, column=12, value=elimination_data)

    merged_cell = [
        f'C{row_value}:E{row_value}',
        f'F{row_value}:H{row_value}',
        f'I{row_value}:K{row_value}',
    ]

    for item in merged_cell:
        await set_merge_cells(worksheet, merged_cell=item)

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    row_dimensions = [
        [f'{row_value}', '105'],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 11, 'center', 'center'),
        (f'O{row_value}:Q{row_value}', 'Times New Roman', 11, 'center', 'center')
    ]

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=row_value, column=12, value=elimination_data)

    for item_range in report_font:
        await set_report_font(
            worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1]
        )

    for item_range in report_font:
        await set_act_alignment(
            worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4]
        )

    await set_automatic_row_dimensions(
        worksheet, row_number=row_value, row_value=violation_values
    )

    return row_value


async def set_act_photographic_materials_values(worksheet: Worksheet, row_value: int):
    """

    :param row_value:
    :param worksheet:
    """

    values = [
        {'coordinate': 'K59', 'value': 'Приложение 1',
         'row': f'{59 + row_value}', 'column': '12'},
        {'coordinate': 'K60', 'value': 'к 1 части Акта-Предписания',
         'row': f'{60 + row_value}', 'column': '12'},
        {'coordinate': 'B62', 'value': 'Фото 1',
         'row': f'{62 + row_value}', 'column': '2'},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_act_photographic_materials_values {repr(err)}")
            continue

    return True


async def set_act_photographic_materials(worksheet: Worksheet, img_params: dict) -> bool:
    """Вставка фото нарушения на страницу акта

    :param img_params:
    :param worksheet:
    :return:
    """

    if not img_params:
        logger.error(f'set_act_photographic_materials not found dict img_params {img_params = }')
        return False

    photo_full_name = img_params.get("photo_full_name", None)
    if not photo_full_name:
        logger.error(
            f'set_act_photographic_materials not found param img_params["photo_full_name"] {photo_full_name = }')
        return False

    if not os.path.isfile(photo_full_name):
        logger.error(f'set_act_photographic_materials photo not found {photo_full_name}')
        return False

    img: Image = Image(photo_full_name)
    img = await image_preparation(img, img_params)
    await insert_images(worksheet, img=img)

    return True


async def set_single_violation(worksheet: Worksheet, violation_values):
    """Заполнение акта из единственного пункта

    :param
    """

    main_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_mainlocation',
        post_id=violation_values.main_location_id
    )
    sub_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_sublocation',
        post_id=violation_values.sub_location_id
    )

    title: str = f"{main_location['title']} ({sub_location['title']})"

    worksheet.cell(row=27, column=2, value=title)
    worksheet.cell(row=28, column=2, value=1)

    normative_document: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=28, column=3, value=violation_values.description)
        # worksheet.cell(row=28, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=28, column=3, value=normative_document['title'])
        # worksheet.cell(row=28, column=15, value=normative_document['title'])

    worksheet.cell(row=28, column=6, value=normative_document['normative'])
    # worksheet.cell(row=28, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in ['', ' ', '.', '*', '/']:
        worksheet.cell(row=28, column=9, value=violation_values.comment)
        # worksheet.cell(row=28, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=28, column=9, value=normative_document['procedure'])
        # worksheet.cell(row=28, column=17, value=normative_document['procedure'])

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    await set_automatic_row_dimensions(worksheet, row_number=28, row_value=violation_values)

    worksheet.cell(row=28, column=12, value=elimination_data)


async def set_value_title(worksheet: Worksheet, violation_values: DataFrame, row_number: int) -> bool:
    """

    :param row_number:
    :param violation_values:
    :param worksheet:
    :return:
    """

    row_value = 28 + row_number
    main_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_mainlocation',
        post_id=violation_values.main_location_id
    )
    sub_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_sublocation',
        post_id=violation_values.sub_location_id
    )

    title: str = f"{main_location['short_title']} ({sub_location['title']})"
    worksheet.cell(row=row_value, column=2, value=title)

    merged_cell = [
        f'B{row_value}:L{row_value}',
    ]
    for item in merged_cell:
        await set_merge_cells(
            worksheet, merged_cell=item
        )

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_range_border(
            worksheet, cell_range=item[0], border=item[1]
        )

    row_dimensions = [
        [f'{row_value}', '18'],
    ]
    for item in row_dimensions:
        await set_row_dimensions(
            worksheet, row_number=item[0], height=item[1]
        )

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 14, 'center', 'center'),
    ]
    for item_range in report_font:
        await set_report_font(
            worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1]
        )

    for item_range in report_font:
        await set_act_alignment(
            worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4]
        )

    report_font = [
        [f'B{row_value}:L{row_value}',
         {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ]

    for cell_range in report_font:
        await sets_report_font(
            worksheet, cell_range[0], params=cell_range[1]
        )

    return True


async def set_act_alignment(worksheet, cell_range, horizontal=None, vertical=None):
    """Форматирование ячейки: положение текста в ячейке (лево верх)

    """
    wrap_alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.alignment = wrap_alignment

        except Exception as err:
            logger.error(f"iter {item} cell {cell}")
            logger.error(f"set_mip_alignment {repr(err)}")


async def set_merge_cells(worksheet: Worksheet, merged_cell: str) -> bool:
    """Форматирование: обьединение ячеек

    :param worksheet:
    :param merged_cell:
    :return:
    """
    try:
        worksheet.merge_cells(merged_cell)
        return True
    except Exception as err:
        logger.error(f"set_merge_cells {repr(err)}")
        return False


async def set_report_font(worksheet, cell_range, font_size=14, font_name='Arial') -> bool:
    """Форматирование ячейки: шрифт и размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(size=font_size, name=font_name)
        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"set_report_font {repr(err)}")
            continue
    return True


async def sets_report_font(worksheet, cell_range, params: dict) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(
                color=params.get("color"),
                italic=params.get("italic"),
                size=params.get("font_size"),
                bold=params.get("bold"),
                name=params.get("name"),
                underline=params.get("underline"),
                vertAlign=params.get("vertAlign"),
            )

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"sets_report_font {repr(err)}")
            continue
    return True


async def set_row_dimensions(worksheet: Worksheet, row_number: int or str, height: int):
    """Установление высоты строки

    :param worksheet:
    :param row_number:
    :param height:
    :return:
    """
    try:
        worksheet.row_dimensions[int(row_number)].height = float(height)
    except Exception as err:
        logger.error(f"set_row_dimensions {repr(err)}")


async def set_automatic_row_dimensions(worksheet: Worksheet, row_number: int, row_value) -> bool:
    """Автоматическое установление высоты строки по тексту

    :param row_value:
    :param worksheet:
    :param row_number:
    :return:
    """

    if not row_value:
        return False

    normative_documents: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=row_value.normative_documents_id)

    title = normative_documents.get('title', None)
    normative = normative_documents.get('normative', None)
    procedure = normative_documents.get('procedure', None)
    comment = row_value.comment

    if not normative:
        logger.error(f"No normative found if {row_value =  }")
    if not procedure:
        logger.error(f"No procedure found if {row_value =  }")

    if procedure:
        comment = row_value.comment if len(row_value.comment) < len(procedure) else procedure

    item_list: list = []
    try:
        list_val = [row_value.description, title, normative, comment]
        for item in list_val:
            if not isinstance(item, str): continue

            lines = min(len(item.split("\n\n")) - 1, 1)
            height = max(len(item) / 26 + lines, 1.5) * 16
            item_list.append(height)

        max_height = round(max(item_list), 2) + 10

        if max_height <= 60:
            max_height = 60

        dim = worksheet.row_dimensions[row_number]
        dim.height = max_height

        logger.debug(f"row_number {row_number} max_height {max_height}")
        return True

    except Exception as err:
        logger.error(f"Error row {row_number} set_automatic_row_dimensions {repr(err)}")
        return False


async def set_border(worksheet):
    """Форматирование ячейки: все границы ячейки

    """
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            try:
                cell.border = thin_border
            except Exception as err:
                logger.error(f"set_border {repr(err)}")


async def set_range_border(worksheet, cell_range, border=None):
    """Форматирование ячейки: все границы ячейки

    border - только нижняя граница
    """

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    if border:
        thin_border = Border(bottom=Side(style='thin'))

    rows = worksheet[cell_range]

    try:
        for row in rows:
            for cell in row:
                try:
                    cell.border = thin_border

                except Exception as err:
                    logger.error(f"set_border {repr(err)}")
                    continue

    except Exception as err:
        logger.error(f"set_border {repr(err)}")
        return False

async def image_preparation(img: Image, img_params: dict):
    """Подготовка изображения перед вставкой на страницу
    """

    height = img_params.get("height")
    scale = img_params.get("scale")
    width = img_params.get("width")
    anchor = img_params.get("anchor")
    column = img_params.get("column")
    row = img_params.get("row")

    # высота изображения
    if height:
        img.height = height

    # ширина изображения
    if width:
        img.width = width

    # изменение пропорций изображения
    if scale:
        scale = img.height / max(img.height, img.width)
        img.width = img.width * scale

    # прикрепление изображение по адресу str(column) + str(row)
    if anchor:
        img.anchor = str(column) + str(row)

    return img


async def insert_images(worksheet: Worksheet, img: Image) -> bool:
    """Вставка изображения на лист worksheet*,
    :param img: файл изображения
    :param worksheet: Worksheet - страница документа для вставки изображения
    :return:
    """

    try:
        worksheet.add_image(img)
        return True

    except Exception as err:
        logger.error(f"insert_images {repr(err)}")
        return False
