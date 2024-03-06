from loader import logger

from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet


async def format_act_prescription_sheet(worksheet: Worksheet):
    """Пошаговое форматирование страницы
    """
    await set_act_page_setup(worksheet)

    for item in ACT_RANGE_COLUMNS:
        await set_column_dimensions(worksheet, column=item[0], width=item[1])

    for merged_cell in ACT_MERGED_CELLS:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    for item in ACT_CELL_RANGES:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    for item in ACT_ROW_DIMENSIONS:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item_range in ACT_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in ACT_CELL_RANGES_BASIC_ALIGNMENT:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    for cell_range in ACT_CELL_RANGES_SET_REPORT_FONT:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])


async def format_act_footer_prescription_sheet(worksheet: Worksheet, row_number: int) -> bool:
    """Пошаговое форматирование страницы

    :param row_number:
    :param worksheet:
    :return: bool
    """
    await set_act_page_setup(worksheet)

    ACT_FOOTER_MERGED_CELLS = [
        f'B{30 + row_number}:L{30 + row_number}', f'B{31 + row_number}:F{31 + row_number}',
        f'G{31 + row_number}:L{31 + row_number}', f'B{32 + row_number}:F{32 + row_number}',
        f'G{32 + row_number}:L{32 + row_number}', f'B{33 + row_number}:L{33 + row_number}',
        f'B{35 + row_number}:L{35 + row_number}', f'B{36 + row_number}:L{36 + row_number}',
        f'B{38 + row_number}:L{38 + row_number}', f'B{39 + row_number}:L{39 + row_number}',
        f'B{40 + row_number}:L{40 + row_number}', f'B{42 + row_number}:F{42 + row_number}',
        f'B{43 + row_number}:F{43 + row_number}', f'H{43 + row_number}:I{43 + row_number}',
        f'K{43 + row_number}:L{43 + row_number}', f'B{44 + row_number}:F{44 + row_number}',
        f'H{44 + row_number}:I{44 + row_number}', f'K{44 + row_number}:L{44 + row_number}',
        f'K{46 + row_number}:L{46 + row_number}', f'K{47 + row_number}:L{47 + row_number}',
        f'B{49 + row_number}:L{49 + row_number}', f'B{50 + row_number}:L{50 + row_number}',
        f'B{51 + row_number}:L{51 + row_number}', f'K{52 + row_number}:L{52 + row_number}',
        f'K{53 + row_number}:L{53 + row_number}', f'K{54 + row_number}:L{54 + row_number}',
        f'K{55 + row_number}:L{55 + row_number}',
    ]
    for merged_cell in ACT_FOOTER_MERGED_CELLS:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    ACT_FOOTER_CELL_RANGES = [
        (f'B{35 + row_number}:L{35 + row_number}', False),
        (f'B{36 + row_number}:L{36 + row_number}', False),
        (f'B{38 + row_number}:L{38 + row_number}', False),
        (f'B{39 + row_number}:L{39 + row_number}', False),
        (f'B{40 + row_number}:L{40 + row_number}', False),
        (f'B{43 + row_number}:F{43 + row_number}', True),
        (f'H{43 + row_number}:I{43 + row_number}', True),
        (f'K{43 + row_number}:L{43 + row_number}', True),
        (f'K{46 + row_number}:L{46 + row_number}', True),
        (f'B{50 + row_number}:L{50 + row_number}', True),
        (f'K{52 + row_number}:L{52 + row_number}', True),
        (f'K{54 + row_number}:L{54 + row_number}', True),
    ]
    for item in ACT_FOOTER_CELL_RANGES:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    ACT_FOOTER_ROW_DIMENSIONS = [
        [f'{29 + row_number}', '5.5'], [f'{30 + row_number}', '45'], [f'{31 + row_number}', '18'],
        [f'{32 + row_number}', '18'], [f'{33 + row_number}', '30'], [f'{34 + row_number}', '5.5'],
        [f'{35 + row_number}', '18'], [f'{36 + row_number}', '32'], [f'{37 + row_number}', '5.5'],
        [f'{38 + row_number}', '20'], [f'{39 + row_number}', '18'], [f'{40 + row_number}', '30'],
        [f'{41 + row_number}', '18'], [f'{42 + row_number}', '18'], [f'{43 + row_number}', '30'],
        [f'{44 + row_number}', '18'], [f'{45 + row_number}', '18'], [f'{46 + row_number}', '18'],
        [f'{47 + row_number}', '18'], [f'{48 + row_number}', '18'], [f'{49 + row_number}', '18'],
        [f'{50 + row_number}', '30'], [f'{51 + row_number}', '40'], [f'{52 + row_number}', '18'],
        [f'{53 + row_number}', '18'], [f'{54 + row_number}', '18'], [f'{55 + row_number}', '18'],
    ]
    for item in ACT_FOOTER_ROW_DIMENSIONS:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT = [
        (f'B{30 + row_number}:L{30 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{31 + row_number}:L{31 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{32 + row_number}:L{32 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{33 + row_number}:L{33 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{35 + row_number}:L{35 + row_number}', 'Times New Roman', 14, 'left', 'center'),
        (f'B{36 + row_number}:L{36 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{38 + row_number}:L{38 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{40 + row_number}:L{40 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{42 + row_number}:L{42 + row_number}', 'Times New Roman', 14, 'center', 'center'),
        (f'B{43 + row_number}:L{43 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{44 + row_number}:L{44 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{46 + row_number}:L{46 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{47 + row_number}:L{47 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{49 + row_number}:L{49 + row_number}', 'Times New Roman', 12, 'left', 'center'),
        (f'B{51 + row_number}:L{51 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{53 + row_number}:L{53 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{55 + row_number}:L{55 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{59 + row_number}:L{59 + row_number}', 'Times New Roman', 10, 'right', 'center'),
        (f'B{60 + row_number}:L{60 + row_number}', 'Times New Roman', 10, 'right', 'center'),
    ]
    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    print_area = f'$A$1:M{56 + row_number}'
    break_line = 58 + row_number
    await set_act_page_after_footer_setup(worksheet, print_area, break_line)

    return True


async def format_act_photo_header(worksheet: Worksheet, row_number: int) -> bool:
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet
    :return: bool
    """
    merged_cells = [
        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', "18"],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "12", "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    return True


async def format_act_photo_description(worksheet: Worksheet, row_number: int):
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet
    :return:
    """
    merged_cells = [

        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', "166"],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "9", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    # background_color = [
    #     ["C50:H50", "FFFFC000"]
    # ]
    # for item in background_color:
    #     await set_report_background_color(worksheet, item[0], rgb=item[1])


async def set_report_font(worksheet, cell_range, font_size=14, font_name='Arial') -> bool:
    """Форматирование ячейки: шрифт и размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(size=font_size, name=font_name)

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"set_report_font {repr(err)}")
            continue
    return True


async def sets_report_font(worksheet, cell_range, params: dict) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(
                color=params.get("color"),
                italic=params.get("italic"),
                size=params.get("font_size"),
                bold=params.get("bold"),
                name=params.get("name"),
                underline=params.get("underline"),
                vertAlign=params.get("vertAlign"),
            )

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"sets_report_font {repr(err)}")
            continue
    return True


async def set_column_dimensions(worksheet, column, width):
    """Форматирование: ширина столбцов

    :param worksheet:
    :param column:
    :param width:
    :return:
    """
    try:
        worksheet.column_dimensions[column].width = width

    except Exception as err:
        logger.error(f"set_column_widths {repr(err)}")


async def set_range_border(worksheet, cell_range, border=None):
    """Форматирование ячейки: все границы ячейки

    border - только нижняя граница
    """

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    if border:
        thin_border = Border(bottom=Side(style='thin'))

    rows = worksheet[cell_range]

    try:
        for row in rows:
            for cell in row:
                try:
                    cell.border = thin_border

                except Exception as err:
                    logger.error(f"set_border {repr(err)}")
                    continue

    except Exception as err:
        logger.error(f"set_border {repr(err)}")
        return False


async def set_merge_cells(worksheet: Worksheet, merged_cell: str) -> bool:
    """Форматирование: обьединение ячеек

    :param worksheet:
    :param merged_cell:
    :return:
    """
    try:
        worksheet.merge_cells(merged_cell)
        return True

    except Exception as err:
        logger.error(f"set_merge_cells {repr(err)}")
        return False


async def set_row_dimensions(worksheet: Worksheet, row_number: int or str, height: int or str):
    """Установление высоты строки

    :param worksheet:
    :param row_number:
    :param height:
    :return:
    """
    try:
        worksheet.row_dimensions[int(row_number)].height = float(height)

    except Exception as err:
        logger.error(f"set_row_dimensions {repr(err)}")


async def set_act_alignment(worksheet, cell_range, horizontal=None, vertical=None):
    """Форматирование ячейки: положение текста в ячейке (лево верх)

    """
    wrap_alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.alignment = wrap_alignment

        except Exception as err:
            logger.error(f"iter {item} cell {cell}")
            logger.error(f"set_mip_alignment {repr(err)}")


async def set_act_page_setup(worksheet: Worksheet) -> bool:
    """Установка параметров страницы

    :param worksheet:
    :return: bool
    """

    #  https://xlsxwriter.readthedocs.io/page_setup.html
    #  worksheet.print_title_rows = '$2:$3'
    #  worksheet.print_title = '$2:$3'

    # Printer Settings
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

    # Подогнать область печати к определенному кол-у страниц как по вертикали, так и по горизонтали.
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = '0'

    # worksheet.views
    worksheet.print_options.horizontalCentered = True
    worksheet.print_area = '$A$1:M56'

    #  масштабный коэффициент для распечатываемой страницы
    #  worksheet.set_print_scale(75)

    #  worksheet.row_breaks.append(Break(id=53))
    #  worksheet.col_breaks.append(Break(id=13))

    # задаем собственные значения отступов
    top_bottom = 2 / 2.54
    left_right = 2 / 2.54
    worksheet.page_margins = PageMargins(left=left_right, right=left_right, top=top_bottom, bottom=top_bottom)

    worksheet.oddFooter.left.text = "Страница &[Page] из &N"
    worksheet.oddFooter.left.size = 10
    worksheet.oddFooter.left.font = "Arial,Bold"
    worksheet.oddFooter.left.color = "030303"
    worksheet.differentFirst = False
    worksheet.differentOddEven = True

    return True


async def set_act_page_after_footer_setup(worksheet: Worksheet, print_area: str, break_line: int = None) -> bool:
    """Установка параметров страницы

    :param break_line: int
    :param print_area: str
    :param worksheet: Worksheet
    :return: bool
    """

    #  https://xlsxons.horizontalCentered = True
    worksheet.print_area = print_area

    #  масштабный коэффициент для распечатываемой страницы
    # worksheet.set_print_scale(75)

    if break_line:
        worksheet.row_breaks.append(Break(id=break_line))
        # worksheet.col_breaks.append(Break(id=13))

    return True


ACT_RANGE_COLUMNS = [
    ['A', '4'],
    ['B', '5'],
    ['C', '9'],
    ['D', '9'],
    ['E', '9'],
    ['F', '12'],
    ['G', '4'],
    ['H', '9'],
    ['I', '9'],
    ['J', '5'],
    ['K', '9'],
    ['L', '12'],
    ['M', '4'],

    ["O", "27"],
    ["P", "25"],
    ["Q", "23"],
]

ACT_MERGED_CELLS = [

    'B6:L6',
    'B7:L7',

    'B9:L9',
    'B10:L10',
    'B11:L11',
    'B12:L12',

    'B14:L14',
    'B15:L15',
    'B16:L16',
    'B17:L17',

    'B19:L19',
    'B20:L20',

    'B22:L22',
    'B23:L23',

    'C25:E26', 'F25:H26', 'I25:K26', 'L25:L26',

    'B27:L27',
    'C28:E28', 'F28:H28', 'I28:K28', 'L28:L28',
]

ACT_CELL_RANGES = [
    ('B9:L9', True),
    ('B12:L12', True),
    ('B16:L16', True),
    ('B20:L20', True),
    ('B25:E26', False), ('F25:H26', False), ('I25:K26', False), ('L25:L26', False),
    ('B27:L27', False),
    ('B28:E28', False), ('F28:H28', False), ('I28:K28', False), ('L28:L28', False),
]

ACT_ROW_DIMENSIONS = [
    ['2', '18'],
    ['3', '18'],
    ['4', '18'],
    ['5', '18'],
    ['6', '18'],
    ['7', '18'],
    ['8', '5.5'],
    ['9', '25'],
    ['10', '40'],
    ['11', '18'],
    ['12', '34'],
    ['13', '5.5'],
    ['14', '22'],
    ['15', '37'],
    ['16', '25'],
    ['17', '25'],
    ['18', '5.5'],
    ['19', '18'],
    ['20', '39'],
    ['21', '5.5'],
    ['22', '18'],
    ['23', '30'],
    ['24', '5.5'],
    ['25', '19'],
    ['26', '53'],
    ['27', '18'],
    ['28', '105'],

]

ACT_CELL_RANGES_BASIC_ALIGNMENT = [
    ('B6:L6', 'Times New Roman', 14, 'center', 'center'),
    ('B7:L7', 'Times New Roman', 14, 'center', 'center'),

    ('B9:L9', 'Verdana', 14, 'center', 'center'),
    ('B10:L10', 'Times New Roman', 10, 'center', 'center'),
    ('B11:L11', 'Times New Roman', 14, 'left', 'center'),
    ('B12:L12', 'Times New Roman', 14, 'left', 'center'),

    ('B14:L14', 'Times New Roman', 10, 'center', 'center'),
    ('B15:L15', 'Times New Roman', 14, 'center', 'center'),
    ('B16:L16', 'Verdana', 14, 'center', 'center'),
    ('B17:L17', 'Times New Roman', 10, 'center', 'center'),

    ('B19:L19', 'Times New Roman', 14, 'left', 'center'),
    ('B20:L20', 'Times New Roman', 14, 'left', 'center'),

    ('B22:L22', 'Times New Roman', 14, 'center', 'center'),
    ('B23:L23', 'Times New Roman', 14, 'center', 'center'),

    ('B25:L25', 'Times New Roman', 11, 'center', 'center'),
    ('B26:L26', 'Times New Roman', 11, 'center', 'center'),
    ('B27:L27', 'Times New Roman', 14, 'center', 'center'),
    ('B28:L28', 'Times New Roman', 11, 'center', 'center'),
    ('O28:Q28', 'Times New Roman', 11, 'center', 'center'),

]

ACT_CELL_RANGES_SET_REPORT_FONT = [
    ['B6:L6', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B7:L7', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B9:L9', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Verdana'}],
    ['B16:L16', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Verdana'}],
    ['B22:L22', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B27:L27', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
]
