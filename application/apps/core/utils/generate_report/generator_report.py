import asyncio

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from apps.MyBot import bot_send_message
from apps.core.bot.messages.messages import Messages
# from apps.core.utils.generate_report.convert_xlsx_to_pdf import convert_report_to_pdf
from apps.core.utils.generate_report.create_xlsx.create_xlsx import create_xlsx
from apps.core.utils.generate_report.generate_act_prescription.set_act_format_ import (format_act_photo_header,
                                                                                       format_act_photo_description)
from apps.core.utils.generate_report.generate_act_prescription.set_act_value import get_act_headlines_data_values
from apps.core.utils.generate_report.generate_act_prescription.set_act_values import set_act_photographic_materials
from apps.core.utils.generate_report.get_file_list import get_json_file_list
from apps.core.utils.generate_report.get_report_path import get_full_report_name
from apps.core.utils.img_processor.insert_img import insert_images_to_sheet
from apps.core.utils.json_worker.read_json_file import read_json_file
from apps.core.utils.reports_processor.report_worker_utils import get_clean_headers
from loader import logger

MAXIMUM_COLUMN_WIDTH: int = 40
MAXIMUM_ROW_HEIGHT: int = 120


async def create_report_from_other_method(chat_id, full_report_path=None,
                                          file_list=None):
    """Создание отчета xls из данных json
    :param file_list:
    :param full_report_path:
    :type chat_id: int
    :param chat_id:
    :return:
    """

    if full_report_path is None:
        full_report_path = await get_full_report_name(chat_id=chat_id)

    workbook, worksheet = await create_xlsx(chat_id, full_report_path)
    if not worksheet:
        return False

    await format_sheets(worksheet)

    if not file_list:
        file_list = await get_json_file_list(chat_id=chat_id)

    await insert_images_to_sheet(file_list, worksheet)

    workbook.save(full_report_path)


async def create_report(chat_id):
    """Создание отчета xls из данных json
    """

    fill_report_path = await get_full_report_name(chat_id=chat_id)
    if fill_report_path is None:
        logger.warning('error! fill_report_path not found!')
        await bot_send_message(chat_id=chat_id, text=Messages.Error.fill_report_path_not_found)
        return

    file_list = await get_json_file_list(chat_id=chat_id)
    if file_list is None:
        logger.warning('error! file_list not found!')
        await bot_send_message(chat_id=chat_id, text=Messages.Error.file_list_not_found)
        return

    dataframe = await create_dataframe(file_list=file_list)
    if dataframe is None:
        logger.warning('error! dataframe not found!')
        await bot_send_message(chat_id=chat_id, text=Messages.Error.dataframe_not_found)
        return

    workbook, worksheet = await create_xlsx(chat_id, fill_report_path)
    if not worksheet:
        return False

    await format_sheets(worksheet)

    await insert_images_to_sheet(file_list, worksheet)

    workbook.save(fill_report_path)

    # await convert_report_to_pdf(chat_id=chat_id, path=fill_report_path)


async def create_dataframe_from_data(data: list) -> DataFrame:
    """Создание dataframe из списка файлов file_list

    :param data: list с данными
    :return: data or None
    """

    column_list = [
        "violation_id",
        "main_category",
        "category",
        "violation_category",
        "general_contractor",
        "description",
        "comment",
        "incident_level",
        "elimination_time",
        "act_required",
        "coordinates",
    ]
    try:
        dataframe = DataFrame(data, columns=column_list)
        return dataframe
    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def create_dataframe(file_list: list) -> DataFrame:
    data = [
        {
            "violation_id": "id записи",
            "main_category": "Основное направление",
            "category": "Категория нарушения",
            "violation_category": "Категория нарушений",
            "general_contractor": "Подрядная организация",
            "description": "Описание нарушения",
            "comment": "Комментарий",
            "incident_level": "Уровень происшествия",
            "elimination_time": "Дней на устранение",
            "act_required": "Оформление акта",
            "coordinates": "Координаты",
        }
    ]

    for file in file_list:
        data.append(await read_json_file(file))

    column_list = [
        "violation_id",
        "main_category",
        "category",
        "violation_category",
        "general_contractor",
        "description",
        "comment",
        "incident_level",
        "elimination_time",
        "act_required",
        "coordinates",
    ]

    try:
        dataframe = DataFrame(data, columns=column_list)
        return dataframe
    except Exception as err:
        logger.error(F"create_dataframe {repr(err)}")
        return None


async def anchor_photo(dataframe, row_number, workbook, worksheet, full_act_path, act_date=None):
    """Вставка фото в ячейку с учетом смещения. Возвращает область печати с учетом фото

    """
    photo_row: int = 63 + row_number
    row_num: int = 0
    img_params: dict = {"height": 220,
                        "anchor": True,
                        "scale": True, }

    row_act_header: int = photo_row - 1
    row_act_photo: int = photo_row
    img_params["row_header"] = row_act_header
    img_params["row_description"] = row_act_photo
    img_params["row"] = row_act_photo

    table_name: str = 'core_violations'
    clean_headers: list = await get_clean_headers(table_name=table_name)

    for num_data, violation_data in enumerate(dataframe.itertuples(index=False), start=1):
        # for num_data, violation_data in enumerate(df.to_dict('index'), start=1):

        violation_dict = dict(zip(clean_headers, violation_data))

        # img_params["photo_full_name"] = str (Udocan_media_path) + "HSE" + os.sep + str(violation_dict.get('photo', None))
        # img_params["photo_full_name"] = await get_photo_full_name(
        #     media_path=await get_directory_name(Udocan_media_path, "HSE"),
        #     photo_path=await get_image_name(violation_dict.get('photo', None))
        # )
        img_params["photo_full_name"] = violation_dict.get('photo_full_name', None)
        if num_data % 2 != 0:
            img_params["column"] = 'B'
            img_params["column_img"] = 2
            img_params["column_description"] = 6
            img_params["description"] = violation_dict.get('description', None)

            if num_data != 1:
                img_params["row_header"] = img_params["row_header"] + 2
                img_params["row_description"] = img_params["row_description"] + 2
                img_params["row"] = img_params["row"] + 2
                await format_act_photo_header(worksheet, img_params["row_header"])
                await format_act_photo_description(worksheet, img_params["row_description"])
                workbook.save(full_act_path)
                photo_row += 2
        else:
            img_params["column"] = 'H'
            img_params["column_img"] = 8
            img_params["column_description"] = 12
            img_params["description"] = violation_dict.get('description', None)

        is_anchor: bool = False
        try:
            is_anchor = await set_act_photographic_materials(worksheet, img_params)

        except Exception as err:
            logger.error(f"set_act_photographic_materials {num_data= } {row_num=} {repr(err)} ")

        if is_anchor:
            workbook.save(full_act_path)
            # set header
            worksheet.cell(row=img_params["row_header"],
                           column=img_params["column_img"],
                           value=f'Фото {num_data}')
            workbook.save(full_act_path)

            # set description
            worksheet.cell(row=img_params["row_description"],
                           column=img_params["column_description"],
                           value=img_params["description"])
            workbook.save(full_act_path)

    print_area = f'$A$1:M{photo_row + row_num + 1}'

    return print_area


async def format_sheets(worksheet: Worksheet) -> Worksheet:
    """Пошаговое форматирование страницы.
    Возвращает измененную страницу.
    """
    await set_border(worksheet)
    await set_alignment(worksheet)
    await set_font(worksheet)
    await set_column_widths(worksheet)
    await set_row_height(worksheet)
    return worksheet


async def set_border(worksheet):
    """Форматирование ячейки: все границы ячейки

    """
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            try:
                cell.border = thin_border
            except Exception as err:
                logger.error(f"set_border {repr(err)}")


async def set_alignment(worksheet: Worksheet):
    """Форматирование ячейки: положение текста в ячейке (лево верх)
    """
    wrap_alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    for row in worksheet.iter_rows():
        for cell in row:
            try:
                cell.alignment = wrap_alignment
            except Exception as err:
                logger.error(f"set_alignment {repr(err)}")


async def set_font(worksheet: Worksheet) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    for row in worksheet.iter_rows():
        for cell in row:
            try:
                cell.font = Font(size=14)
            except Exception as err:
                logger.error(f"sets_report_font {repr(err)}")
                continue
    return True


async def set_column_widths(worksheet: Worksheet):
    """Форматирование ячейки: ширина столбца

    """

    for column_cells in worksheet.columns:
        # максимальная ширина столбца
        column_length = max(len(_as_text(cell.value)) for cell in column_cells)

        if column_length < MAXIMUM_COLUMN_WIDTH:
            new_column_length = column_length
        else:
            new_column_length = MAXIMUM_COLUMN_WIDTH

        new_column_letter: int = (openpyxl.utils.get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            try:
                worksheet.column_dimensions[new_column_letter].width = new_column_length + 1
            except Exception as err:
                logger.error(f"set_column_widths {repr(err)}")


def _as_text(value) -> str:
    """Приведение данных к str

    """
    if value is None:
        return ""
    return str(value)


async def set_row_height(worksheet: Worksheet):
    """Форматирование ячейки: высота шрифта
    """
    for ind in range(worksheet.max_row):
        if ind == 0:
            continue
        try:
            worksheet.row_dimensions[ind + 1].height = MAXIMUM_ROW_HEIGHT
        except Exception as err:
            logger.error(F"set_row_height {repr(err)}")


async def test():
    chat_id = 373084462
    await get_act_headlines_data_values(chat_id)


if __name__ == "__main__":
    asyncio.run(test())
