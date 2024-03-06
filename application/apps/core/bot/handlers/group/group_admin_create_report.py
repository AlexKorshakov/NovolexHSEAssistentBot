from __future__ import annotations

import os

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from apps.MyBot import bot_send_message
from apps.core.bot.handlers.generate.generate_act_prescription_support import (
    set_act_body_values,
    set_act_page_after_footer_setup,
    set_act_violation_values,
    set_act_photographic_materials_values,
    set_act_footer_values,
    set_act_footer_footer_values,
    set_act_header_values,
    get_act_headlines_data_values,
)
from apps.core.bot.handlers.group.group_write_data import db_get_clean_headers
from apps.core.bot.handlers.group.group_support_path import (get_dop_photo_full_filename,
                                                             get_image_name,
                                                             get_photo_full_name)
from apps.core.bot.messages.messages import Messages
from loader import logger


async def create_act_prescription(chat_id: int, act_number: int | str, dataframe: DataFrame, full_act_path: str,
                                  act_date: str = None, qr_img_insert: bool = False) -> bool:
    """Формирование Акта-предписания из dataframe

    :param qr_img_insert:
    :param act_number: int
    :param act_date:
    :param full_act_path:
    :param chat_id: int
    :param dataframe: DataFrame
    """
    act_number = str(act_number)
    if not full_act_path:
        logger.warning(Messages.Error.fill_report_path_not_found)
        return False

    workbook, worksheet = await create_xlsx(chat_id, full_act_path)
    if not worksheet:
        return False

    await format_act_prescription_sheet(worksheet, act_num=act_number)

    await set_act_body_values(worksheet)
    workbook.save(full_act_path)

    await insert_service_image(worksheet, chat_id=chat_id, service_image_name=f'company_{chat_id}')

    if qr_img_insert:
        qr_img_params: dict = {
            "height": 150,
            "width": 150,
            "anchor": True,
            "column": 'K',
            "column_img": 11,
            "row": 1,
        }
        service_image_name: str = await get_image_name(
                'data_file', act_date, 'reports', f'qr_act_nom_{chat_id}_{act_number}'
        )

        await insert_service_image(
                worksheet, chat_id=chat_id, service_image_name=service_image_name, img_params=qr_img_params
        )

    headlines_data_values: dict = await get_act_headlines_data_values(
            chat_id=chat_id, dataframe=dataframe, act_date=act_date, act_number=act_number
    )
    await set_act_header_values(
            worksheet, headlines_data=headlines_data_values
    )
    workbook.save(full_act_path)

    try:
        row_number = await set_act_violation_values(worksheet, dataframe, workbook, full_act_path)

    except Exception as err:
        await bot_send_message(chat_id=chat_id, text=f'Messages.Error.file_not_found {repr(err)} ')
        logger.error(f'create_act_prescription: set_act_violation_values error: {repr(err)}')
        return False

    await format_act_footer_prescription_sheet(worksheet, row_number, act_number=act_number)
    await set_act_footer_values(worksheet, row_number)
    await set_act_footer_footer_values(worksheet, row_number)
    workbook.save(full_act_path)

    img_params: dict = {
        "height": 62,
        "width": 130,
        "anchor": True,
        "column": 'H',
        "column_img": 8,
        "row": 42 + row_number,
    }

    await insert_service_image(
            worksheet, chat_id=chat_id, service_image_name=f'signature_{chat_id}', img_params=img_params
    )

    await set_act_photographic_materials_values(worksheet, row_number)
    # await format_act_photographic(worksheet, row_number)
    await format_act_photo_header(worksheet, row_number=62 + row_number)
    await format_act_photo_description(worksheet, row_number=63 + row_number)
    workbook.save(full_act_path)

    print_area, row_number = await anchor_photo(
            dataframe=dataframe, row_number=row_number, workbook=workbook, worksheet=worksheet, full_act_path=full_act_path
    )

    await set_act_page_after_footer_setup(worksheet, print_area)
    workbook.save(full_act_path)

    row_number = row_number + 3
    print_area, row_number = await anchor_dop_photo(
            chat_id=chat_id, dataframe=dataframe, start_row_number=row_number, workbook=workbook, worksheet=worksheet,
            full_act_path=full_act_path,
    )
    await set_act_page_after_footer_setup(worksheet, print_area)
    workbook.save(full_act_path)

    return True


async def insert_service_image(worksheet: Worksheet, *, chat_id: int = None, service_image_name: str = 'Logo',
                               img_params: dict = None) -> bool:
    """Вставка изображений в файл

    :param service_image_name: str - имя файла для обработки
    :param chat_id: int - id пользователя (папки) где находится logo
    :param img_params: dict параметры вставки
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """
    photo_full_name: str = await get_photo_full_name(chat_id=chat_id, image_name=f"{service_image_name}.jpg")

    # if chat_id:
    #     photo_full_name: str = await get_image_name(Novolex_media_path, "HSE",
    #     str(chat_id), f"{service_image_name}.jpg")

    if not os.path.isfile(photo_full_name):
        logger.error("service image not found")
        photo_full_name: str = await get_photo_full_name(chat_id=chat_id, image_name="Logo.jpg")

    if not img_params:
        img_params: dict = {
            'photo_full_name': photo_full_name,
            "height": 90,
            "width": 230,
            "anchor": True,
            "column": 'B',
            "column_img": 2,
            "row": 2,
        }

    if not os.path.isfile(photo_full_name):
        logger.error("service image not found")
        return False

    img: Image = Image(photo_full_name)
    img = await image_preparation(img, img_params)

    result = await insert_images(worksheet, img=img)
    return bool(result)


async def anchor_dop_photo(chat_id, dataframe: DataFrame, start_row_number: int, workbook: Workbook,
                           worksheet: Worksheet, full_act_path: str):
    """Добавление дополнительных фотоматериалов"""
    img_params: list = await get_img_params(
            user_id=chat_id, dataframe=dataframe, start_row_num=start_row_number
    )
    photo_row: int = start_row_number
    for photo_num, img_param in enumerate(img_params, start=1):
        try:
            await set_act_photographic_materials(worksheet, img_param)
            await set_row_dimensions(worksheet, row_number=img_param['row'], height=242)
            await set_act_photographic_materials_description(worksheet, img_param)
            photo_row: int = img_param.get('row')
            workbook.save(full_act_path)

        except Exception as err:
            logger.error(f"set_act_photographic_materials {photo_num= } {repr(err)} ")
            continue

    print_area = f'$A$1:M{photo_row + 1}'
    return print_area, photo_row + 1


async def get_img_params(user_id: str, dataframe: DataFrame, start_row_num: int) -> list:
    """

    :return:
    """
    group_photo_list: list = await get_group_photo_list(user_id=user_id, dataframe=dataframe)
    photo_row: int = start_row_num

    img_params_list: list = []
    for num_data, media_file_dict in enumerate(group_photo_list, start=1):
        img_params: dict = {"height": 320, "anchor": True, "scale": True, "row_header": photo_row - 1,
                            "row_description": photo_row,
                            'photo_full_name': media_file_dict.get('photo_full_name', None),
                            "description": media_file_dict.get('description', None)}

        if num_data % 2 != 0:
            img_params["column"] = 'B'
            img_params["column_img"] = 2
            # img_params["column_description"] = 6
            img_params["row"] = photo_row

            if num_data != 1:
                # img_params["row_header"] = photo_row - 1
                img_params["row_description"] = img_params["row_description"] + 2
        else:
            img_params["column"] = 'H'
            img_params["column_img"] = 8
            # img_params["column_description"] = 12
            # img_params["row_header"] = photo_row - 1
            img_params["row"] = photo_row
            photo_row += 2

        img_params_list.append(img_params)

    return img_params_list


async def get_group_photo_list(user_id: str, dataframe: DataFrame) -> list:
    """

    :return:
    """
    clean_headers: list = await db_get_clean_headers(table_name='core_violations')

    group_list: list = []
    for violation_num, violation_data in enumerate(dataframe.itertuples(index=False), start=1):
        violation_dict: dict = dict(zip(clean_headers, violation_data))
        if not violation_dict.get('media_group', None): continue
        media_group_list: list = violation_dict.get('media_group', None).split(':::')

        for num_in_group, item in enumerate(media_group_list, start=1):
            photo_full_name: str = await get_dop_photo_full_filename(user_id=user_id, name=item)

            if not os.path.isfile(photo_full_name):
                continue

            group_list.append(
                    {
                        'description': f'Доп.фото №{num_in_group} к нарушению №{violation_num}',
                        'photo_full_name': photo_full_name,
                    }
            )
    return group_list


async def create_xlsx(chat_id: int | str, full_act_path: str) -> (Workbook, Worksheet):
    """Создание файла xlsx с нуля

    :param: chat_id
    :param: full_act_path

    """

    is_created: bool = await create_new_xlsx(report_file=full_act_path)
    if is_created is None:
        logger.warning(Messages.Error.workbook_not_create)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_create)
        return None, None

    workbook: Workbook = await get_workbook(fill_report_path=full_act_path)
    if workbook is None:
        logger.warning(Messages.Error.workbook_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_found)
        return None, None

    worksheet: Worksheet = await get_worksheet(workbook, index=0)
    if worksheet is None:
        logger.warning(Messages.Error.worksheet_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.worksheet_not_found)
        return None, None

    return workbook, worksheet


async def create_new_xlsx(report_file: str) -> bool:
    """Создание xlsx
    """
    try:
        wb = openpyxl.Workbook()
        wb.save(report_file)
        return True

    except Exception as err:
        logger.error(F"set_border {repr(err)}")
        return False


async def get_worksheet(wb: Workbook, index: int = 0) -> Worksheet | None:
    """Получение Страницы из документа по индексу
    :param wb: Workbook - книга xls
    :param index: int - индекс листа
    :return: worksheet or None
    """
    try:
        worksheet: Worksheet = wb.worksheets[index]
        return worksheet

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def get_workbook(fill_report_path: str) -> Workbook | None:
    """Открыть и загрузить Workbook
    :param fill_report_path: str полный путь к файлу
    :return: Workbook or None
    """
    try:
        workbook: Workbook = openpyxl.load_workbook(fill_report_path)
        return workbook

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def anchor_photo(dataframe, row_number: int, workbook: Workbook, worksheet: Worksheet, full_act_path: str):
    """Вставка фото в ячейку с учетом смещения. Возвращает область печати с учетом фото

    """
    photo_row: int = 63 + row_number
    row_num: int = 0
    img_params: dict = {"height": 220,
                        "anchor": True,
                        "scale": True, }

    row_act_header: int = photo_row - 1
    row_act_photo: int = photo_row
    img_params["row_header"] = row_act_header
    img_params["row_description"] = row_act_photo
    img_params["row"] = row_act_photo

    table_name: str = 'core_violations'
    clean_headers: list = await db_get_clean_headers(table_name=table_name)

    for num_data, violation_data in enumerate(dataframe.itertuples(index=False), start=1):
        # for num_data, violation_data in enumerate(df.to_dict('index'), start=1):

        violation_dict = dict(zip(clean_headers, violation_data))

        img_params["photo_full_name"] = violation_dict.get('photo_full_name', None)
        if num_data % 2 != 0:
            img_params["column"] = 'B'
            img_params["column_img"] = 2
            img_params["column_description"] = 6
            img_params["description"] = violation_dict.get('description', None)

            if num_data != 1:
                img_params["row_header"] = img_params["row_header"] + 2
                img_params["row_description"] = img_params["row_description"] + 2
                img_params["row"] = img_params["row"] + 2
                await format_act_photo_header(worksheet, img_params["row_header"])
                await format_act_photo_description(worksheet, img_params["row_description"])
                workbook.save(full_act_path)
                photo_row += 2
        else:
            img_params["column"] = 'H'
            img_params["column_img"] = 8
            img_params["column_description"] = 12
            img_params["description"] = violation_dict.get('description', None)

        is_anchor: bool = False
        try:
            is_anchor = await set_act_photographic_materials(worksheet, img_params)

        except Exception as err:
            logger.error(f"set_act_photographic_materials {num_data= } {row_num=} {repr(err)} ")

        if is_anchor:
            workbook.save(full_act_path)
            # set header
            worksheet.cell(row=img_params["row_header"],
                           column=img_params["column_img"],
                           value=f'Фото {num_data} к пункту {num_data} ')
            workbook.save(full_act_path)

            # set description
            worksheet.cell(row=img_params["row_description"],
                           column=img_params["column_description"],
                           value=img_params["description"])
            workbook.save(full_act_path)

    print_area = f'$A$1:M{photo_row + row_num + 1}'

    return print_area, photo_row + row_num + 1


async def set_act_photographic_materials(worksheet: Worksheet, img_params: dict) -> bool:
    """Вставка фото нарушения на страницу акта

    :param img_params:
    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    if not img_params:
        logger.error(f'set_act_photographic_materials not found dict img_params {img_params = }')
        return False

    photo_full_name = img_params.get("photo_full_name", None)
    if not photo_full_name:
        logger.error(
                f'set_act_photographic_materials not found param img_params["photo_full_name"] {photo_full_name = }')
        return False

    if not os.path.isfile(photo_full_name):
        logger.error(f'set_act_photographic_materials photo not found {photo_full_name}')
        return False

    img: Image = Image(photo_full_name)
    img = await image_preparation(img, img_params)
    await insert_images(worksheet, img=img)

    return True


async def image_preparation(img: Image, img_params: dict) -> Image:
    """Подготовка изображения перед вставкой на страницу

    :param: img - объект изображения
    :param: img_params - параметры изображения
    :return: Image - измененный объект изображения
    """
    height = img_params.get("height")
    scale = img_params.get("scale")
    width = img_params.get("width")
    anchor = img_params.get("anchor")
    column = img_params.get("column")
    row = img_params.get("row")

    # высота изображения
    if height:
        img.height = height

    # ширина изображения
    if width:
        img.width = width

    # изменение пропорций изображения
    if scale:
        scale = img.height / max(img.height, img.width)
        img.width = img.width * scale

    # прикрепление изображение по адресу str(column) + str(row)
    if anchor:
        img.anchor = str(column) + str(row)

    return img


async def insert_images(worksheet: Worksheet, img: Image) -> bool:
    """Вставка изображения на лист worksheet,

    :param img: файл изображения
    :param worksheet: Worksheet - страница документа для вставки изображения
    :return: bool
    """

    try:
        worksheet.add_image(img)
        return True

    except Exception as err:
        logger.error(f"insert_images {repr(err)}")
        return False


async def format_act_photo_header(worksheet: Worksheet, row_number: int) -> bool:
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """
    merged_cells = [
        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', 18],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "12", "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    return True


async def format_act_photo_description(worksheet: Worksheet, row_number: int) -> bool:
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    merged_cells = [

        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', "166"],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "9", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    # background_color = [
    #     ["C50:H50", "FFFFC000"]
    # ]
    # for item in background_color:
    #     await set_report_background_color(worksheet, item[0], rgb=item[1])

    return True


async def format_act_footer_prescription_sheet(worksheet: Worksheet, row_number: int, act_number: int | str) -> bool:
    """Пошаговое форматирование страницы

    :param act_number:
    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """
    await set_act_page_setup(worksheet, act_num=act_number)

    ACT_FOOTER_MERGED_CELLS = [
        f'B{30 + row_number}:L{30 + row_number}', f'B{31 + row_number}:F{31 + row_number}',
        f'G{31 + row_number}:L{31 + row_number}', f'B{32 + row_number}:F{32 + row_number}',
        f'G{32 + row_number}:L{32 + row_number}', f'B{33 + row_number}:L{33 + row_number}',
        f'B{35 + row_number}:L{35 + row_number}', f'B{36 + row_number}:L{36 + row_number}',
        f'B{38 + row_number}:L{38 + row_number}', f'B{39 + row_number}:L{39 + row_number}',
        f'B{40 + row_number}:L{40 + row_number}', f'B{42 + row_number}:F{42 + row_number}',
        f'B{43 + row_number}:F{43 + row_number}', f'H{43 + row_number}:I{43 + row_number}',
        f'K{43 + row_number}:L{43 + row_number}', f'B{44 + row_number}:F{44 + row_number}',
        f'H{44 + row_number}:I{44 + row_number}', f'K{44 + row_number}:L{44 + row_number}',
        f'K{46 + row_number}:L{46 + row_number}', f'K{47 + row_number}:L{47 + row_number}',
        f'B{49 + row_number}:L{49 + row_number}', f'B{50 + row_number}:L{50 + row_number}',
        f'B{51 + row_number}:L{51 + row_number}', f'K{52 + row_number}:L{52 + row_number}',
        f'K{53 + row_number}:L{53 + row_number}', f'K{54 + row_number}:L{54 + row_number}',
        f'K{55 + row_number}:L{55 + row_number}',
    ]
    for merged_cell in ACT_FOOTER_MERGED_CELLS:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    ACT_FOOTER_CELL_RANGES = [
        (f'B{35 + row_number}:L{35 + row_number}', False),
        (f'B{36 + row_number}:L{36 + row_number}', False),
        (f'B{38 + row_number}:L{38 + row_number}', False),
        (f'B{39 + row_number}:L{39 + row_number}', False),
        (f'B{40 + row_number}:L{40 + row_number}', False),
        (f'B{43 + row_number}:F{43 + row_number}', True),
        (f'H{43 + row_number}:I{43 + row_number}', True),
        (f'K{43 + row_number}:L{43 + row_number}', True),
        (f'K{46 + row_number}:L{46 + row_number}', True),
        (f'B{50 + row_number}:L{50 + row_number}', True),
        (f'K{52 + row_number}:L{52 + row_number}', True),
        (f'K{54 + row_number}:L{54 + row_number}', True),
    ]
    for item in ACT_FOOTER_CELL_RANGES:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    ACT_FOOTER_ROW_DIMENSIONS = [
        [f'{29 + row_number}', '5.5'], [f'{30 + row_number}', '45'], [f'{31 + row_number}', '18'],
        [f'{32 + row_number}', '18'], [f'{33 + row_number}', '30'], [f'{34 + row_number}', '5.5'],
        [f'{35 + row_number}', '18'], [f'{36 + row_number}', '32'], [f'{37 + row_number}', '5.5'],
        [f'{38 + row_number}', '20'], [f'{39 + row_number}', '18'], [f'{40 + row_number}', '30'],
        [f'{41 + row_number}', '18'], [f'{42 + row_number}', '18'], [f'{43 + row_number}', '30'],
        [f'{44 + row_number}', '18'], [f'{45 + row_number}', '18'], [f'{46 + row_number}', '18'],
        [f'{47 + row_number}', '18'], [f'{48 + row_number}', '18'], [f'{49 + row_number}', '18'],
        [f'{50 + row_number}', '30'], [f'{51 + row_number}', '40'], [f'{52 + row_number}', '18'],
        [f'{53 + row_number}', '18'], [f'{54 + row_number}', '18'], [f'{55 + row_number}', '18'],
    ]
    for item in ACT_FOOTER_ROW_DIMENSIONS:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT = [
        (f'B{30 + row_number}:L{30 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{31 + row_number}:L{31 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{32 + row_number}:L{32 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{33 + row_number}:L{33 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{35 + row_number}:L{35 + row_number}', 'Times New Roman', 14, 'left', 'center'),
        (f'B{36 + row_number}:L{36 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{38 + row_number}:L{38 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{40 + row_number}:L{40 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{42 + row_number}:L{42 + row_number}', 'Times New Roman', 14, 'center', 'center'),
        (f'B{43 + row_number}:L{43 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{44 + row_number}:L{44 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{46 + row_number}:L{46 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{47 + row_number}:L{47 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{49 + row_number}:L{49 + row_number}', 'Times New Roman', 12, 'left', 'center'),
        (f'B{51 + row_number}:L{51 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{53 + row_number}:L{53 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{55 + row_number}:L{55 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{59 + row_number}:L{59 + row_number}', 'Times New Roman', 10, 'right', 'center'),
        (f'B{60 + row_number}:L{60 + row_number}', 'Times New Roman', 10, 'right', 'center'),
    ]
    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    print_area = f'$A$1:M{56 + row_number}'
    break_line = 58 + row_number
    await set_act_page_after_footer_setup(worksheet, print_area, break_line)

    return True


async def format_act_prescription_sheet(worksheet: Worksheet, act_num: str):
    """Пошаговое форматирование страницы
    """
    await set_act_page_setup(worksheet, act_num)

    for item in ACT_RANGE_COLUMNS:
        await set_column_dimensions(worksheet, column=item[0], width=item[1])

    for merged_cell in ACT_MERGED_CELLS:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    for item in ACT_CELL_RANGES:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    for item in ACT_ROW_DIMENSIONS:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item_range in ACT_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in ACT_CELL_RANGES_BASIC_ALIGNMENT:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    for cell_range in ACT_CELL_RANGES_SET_REPORT_FONT:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])


async def set_column_dimensions(worksheet, column, width):
    """Форматирование: ширина столбцов

    :param worksheet:
    :param column:
    :param width:
    :return:
    """
    try:
        worksheet.column_dimensions[column].width = width
    except Exception as err:
        logger.error(f"set_column_widths {repr(err)}")


async def set_report_font(worksheet, cell_range, font_size=14, font_name='Arial') -> bool:
    """Форматирование ячейки: шрифт и размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(size=font_size, name=font_name)
        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"set_report_font {repr(err)}")
            continue
    return True


async def set_act_alignment(worksheet, cell_range, horizontal=None, vertical=None):
    """Форматирование ячейки: положение текста в ячейке (лево верх)

    """
    wrap_alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.alignment = wrap_alignment

        except Exception as err:
            logger.error(f"iter {item} cell {cell}")
            logger.error(f"set_mip_alignment {repr(err)}")


async def set_merge_cells(worksheet: Worksheet, merged_cell: str) -> bool:
    """Форматирование: обьединение ячеек

    :param worksheet:
    :param merged_cell:
    :return:
    """
    try:
        worksheet.merge_cells(merged_cell)
        return True
    except Exception as err:
        logger.error(f"set_merge_cells {repr(err)}")
        return False


async def set_act_page_setup(worksheet: Worksheet, act_num=None) -> bool:
    """Установка параметров страницы

    :param act_num:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """

    #  https://xlsxwriter.readthedocs.io/page_setup.html
    #  worksheet.print_title_rows = '$2:$3'
    #  worksheet.print_title = '$2:$3'

    # Printer Settings
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

    # Подогнать область печати к определенному кол-у страниц как по вертикали, так и по горизонтали.
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = '0'

    # worksheet.views
    worksheet.print_options.horizontalCentered = True
    worksheet.print_area = '$A$1:M56'

    #  масштабный коэффициент для распечатываемой страницы
    #  worksheet.set_print_scale(75)

    #  worksheet.row_breaks.append(Break(id=53))
    #  worksheet.col_breaks.append(Break(id=13))

    # задаем собственные значения отступов
    top_bottom = 2 / 2.54
    left_right = 2 / 2.54
    worksheet.page_margins = PageMargins(left=left_right, right=left_right, top=top_bottom, bottom=top_bottom)

    worksheet.oddFooter.left.text = "Страница &[Page] из &N"

    if act_num:
        worksheet.oddFooter.left.text = f"Страница &[Page] из &N  акта № {act_num}"

    worksheet.oddFooter.left.size = 10
    worksheet.oddFooter.left.font = "Arial,Bold"
    worksheet.oddFooter.left.color = "030303"
    worksheet.differentFirst = False
    worksheet.differentOddEven = True

    return True


async def sets_report_font(worksheet, cell_range, params: dict) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(
                    color=params.get("color"),
                    italic=params.get("italic"),
                    size=params.get("font_size"),
                    bold=params.get("bold"),
                    name=params.get("name"),
                    underline=params.get("underline"),
                    vertAlign=params.get("vertAlign"),
            )

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"sets_report_font {repr(err)}")
            continue
    return True


async def set_row_dimensions(worksheet: Worksheet, row_number: int | str, height: int | str):
    """Установление высоты строки

    :param worksheet: Worksheet - объект страницы файла
    :param row_number: int | str - номер страницы
    :param height: int - высота
    :return:
    """
    try:
        worksheet.row_dimensions[int(row_number)].height = float(height)
    except Exception as err:
        logger.error(f"set_row_dimensions {repr(err)}")


async def set_range_border(worksheet, cell_range, border=None):
    """Форматирование ячейки: все границы ячейки

    border - только нижняя граница
    """

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    if border:
        thin_border = Border(bottom=Side(style='thin'))

    rows = worksheet[cell_range]

    try:
        for row in rows:
            for cell in row:
                try:
                    cell.border = thin_border

                except Exception as err:
                    logger.error(f"set_border {repr(err)}")
                    continue

    except Exception as err:
        logger.error(f"set_border {repr(err)}")
        return False


async def set_act_photographic_materials_description(worksheet, img_params):
    # set header
    worksheet.cell(
            row=img_params["row_header"], column=img_params["column_img"], value=img_params["description"]
    )


ACT_RANGE_COLUMNS = [
    ['A', '4'],
    ['B', '5'],
    ['C', '9'],
    ['D', '9'],
    ['E', '9'],
    ['F', '12'],
    ['G', '4'],
    ['H', '9'],
    ['I', '9'],
    ['J', '5'],
    ['K', '9'],
    ['L', '12'],
    ['M', '4'],

    ["O", "27"],
    ["P", "25"],
    ["Q", "23"],
]

ACT_MERGED_CELLS = [

    'B6:L6',
    'B7:L7',

    'B9:L9',
    'B10:L10',
    'B11:L11',
    'B12:L12',

    'B14:L14',
    'B15:L15',
    'B16:L16',
    'B17:L17',

    'B19:L19',
    'B20:L20',

    'B22:L22',
    'B23:L23',

    'C25:E26', 'F25:H26', 'I25:K26', 'L25:L26',

    'B27:L27',
    'C28:E28', 'F28:H28', 'I28:K28', 'L28:L28',
]

ACT_CELL_RANGES = [
    ('B9:L9', True),
    ('B12:L12', True),
    ('B16:L16', True),
    ('B20:L20', True),
    ('B25:E26', False), ('F25:H26', False), ('I25:K26', False), ('L25:L26', False),
    ('B27:L27', False),
    ('B28:E28', False), ('F28:H28', False), ('I28:K28', False), ('L28:L28', False),
]

ACT_ROW_DIMENSIONS = [
    ['2', '18'],
    ['3', '18'],
    ['4', '18'],
    ['5', '18'],
    ['6', '18'],
    ['7', '18'],
    ['8', '5.5'],
    ['9', '25'],
    ['10', '40'],
    ['11', '18'],
    ['12', '34'],
    ['13', '5.5'],
    ['14', '22'],
    ['15', '37'],
    ['16', '25'],
    ['17', '25'],
    ['18', '5.5'],
    ['19', '18'],
    ['20', '39'],
    ['21', '5.5'],
    ['22', '18'],
    ['23', '30'],
    ['24', '5.5'],
    ['25', '19'],
    ['26', '53'],
    ['27', '18'],
    ['28', '105'],

]

ACT_CELL_RANGES_BASIC_ALIGNMENT = [
    ('B6:L6', 'Times New Roman', 14, 'center', 'center'),
    ('B7:L7', 'Times New Roman', 14, 'center', 'center'),

    ('B9:L9', 'Verdana', 14, 'center', 'center'),
    ('B10:L10', 'Times New Roman', 10, 'center', 'center'),
    ('B11:L11', 'Times New Roman', 14, 'left', 'center'),
    ('B12:L12', 'Times New Roman', 14, 'left', 'center'),

    ('B14:L14', 'Times New Roman', 10, 'center', 'center'),
    ('B15:L15', 'Times New Roman', 14, 'center', 'center'),
    ('B16:L16', 'Verdana', 14, 'center', 'center'),
    ('B17:L17', 'Times New Roman', 10, 'center', 'center'),

    ('B19:L19', 'Times New Roman', 14, 'left', 'center'),
    ('B20:L20', 'Times New Roman', 14, 'left', 'center'),

    ('B22:L22', 'Times New Roman', 14, 'center', 'center'),
    ('B23:L23', 'Times New Roman', 14, 'center', 'center'),

    ('B25:L25', 'Times New Roman', 11, 'center', 'center'),
    ('B26:L26', 'Times New Roman', 11, 'center', 'center'),
    ('B27:L27', 'Times New Roman', 14, 'center', 'center'),
    ('B28:L28', 'Times New Roman', 11, 'center', 'center'),
    ('O28:Q28', 'Times New Roman', 11, 'center', 'center'),

]

ACT_CELL_RANGES_SET_REPORT_FONT = [
    ['B6:L6', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B7:L7', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B9:L9', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Verdana'}],
    ['B16:L16', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Verdana'}],
    ['B22:L22', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B27:L27', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
]
