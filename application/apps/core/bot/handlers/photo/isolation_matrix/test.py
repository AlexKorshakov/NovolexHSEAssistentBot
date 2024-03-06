from __future__ import annotations

import asyncio
import openpyxl
import os
import shutil
import time
import traceback
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

import pandas
from pandas import DataFrame

from apps.MyBot import bot_send_message
from apps.core.bot.handlers.photo.AmazingQRCodeGenerator import create_qr_code
from apps.core.bot.messages.messages import Messages
from config.config import Udocan_media_path
from loader import logger

SUB_LOCATION: str = '!all_matrix'
VERSION: str = 'v 0.1'


async def create_industrial_equipment_catalog(chat_id: str, industrial_equipment_file_list: list, main_path: str):
    """"""

    fool_file_patch: str = str(Path(main_path, SUB_LOCATION, '!!!Оборудование!База данных оборудования БДО.xlsx'))
    sheet_name: str = 'ЕО'
    equipment_codes_df: DataFrame = pandas.read_excel(fool_file_patch, sheet_name=sheet_name)

    catalog_path: str = str(Path(main_path, '!industrial_equipment_catalog'))
    if not os.path.isdir(catalog_path):
        os.makedirs(catalog_path)

    for number, item in enumerate(industrial_equipment_file_list, start=1):
        # if number == 50: return

        equipment_desc: str = ''
        equipment_desc_df: DataFrame = equipment_codes_df[equipment_codes_df.eq(item.get('matrix_code')).any(1)]
        if await check_dataframe(equipment_desc_df, chat_id):
            # logger.info(f'{equipment_desc_df = }')
            equipment_desc = f" {equipment_desc_df['Наименование ЕО/ТМ'].values[0]}"
            if equipment_desc: equipment_desc = equipment_desc[:55]
            equipment_desc = equipment_desc.split('/')[0]

        equipment_desc = equipment_desc.strip()

        src_file: str = item.get('full_file_path')
        if not os.path.isfile(src_file):
            continue

        # if number == 10:
        #     print()

        suffix = Path(src_file).suffix
        matrix_num: str = '{:004}'.format(number)
        matrix_code: str = item.get('matrix_code')
        if matrix_code in [
            'ZG01-GM-103', "SF01-CV-131", 'ZF01-ТК-101', 'ZF01-ТК-105', 'ST01-SA-102', 'ZF01-SA-133', 'ZF01-SA-132',
            'ZF01-SA-131', 'ZF01-SA-122', 'ZF01-SA-121', 'ZF1-PU-157', 'ZF01-PU-293', 'ZF01-PU-292', 'ZF01-PU-291',
            'ZF1-PU-156', 'ZF1-PU-162', 'ZF1-PU-161', 'ZF01-SA-104', 'ZF01-SA-103', 'ZF01-SA-102', 'ZF01-SA-101',
            'ZF01-SA-104', 'ZF01-ТК-108', 'ZF01-ТК-121', 'ZG01-SA-207', 'ZF01-PU-114', 'ZF01-PU-114', 'ZG01-CY-201',
            'ZG01-CY-104', 'ZG01-CY-204', 'ZG01-CY-203', 'ZF01-PU-113', 'ZF01-PU-112', 'ZG01-CY-105', 'ZF01-PU-111',
            'ZF01-PU-110', 'ZF01-PU-193', 'ZF01-PU-192', 'ZF01-PU-191', 'ZF01-PU-253', 'ZF01-ТК-201', 'ZG01-SA-107',
            'US01-CV-153', 'US01-CV-154', 'US01-CV-155', 'US01-CV-156', 'ZG01-CY-101', 'ZG01-CY-102', 'ZG01-CY-103',
            'ZG01-CY-202'

        ]: continue

        location: str = f"{item.get('glob_location')}"

        await create_main_folder(catalog_path, location)

        matrix_dir: str = f"{matrix_num} {matrix_code} {location} {equipment_desc}"
        await create_sub_folder(catalog_path, matrix_dir)

        destination_file_name: str = f'{matrix_code} {location} {equipment_desc}'
        destination_file: str = str(
            Path(catalog_path, matrix_dir, '03 Матрица изоляции',
                 f'{matrix_num} Матрица изоляции {destination_file_name}{suffix}'))
        if not os.path.isfile(destination_file):
            await save_doc_file(src_file, destination_file)

        save_dir = str(Path(catalog_path, matrix_dir, '88 Прочее'))
        save_name = await save_qr_code(matrix_code, save_dir)

        await create_qr_xls_file(
            chat_id=chat_id, qr_path=save_dir,
            save_name=save_name, matrix_num=matrix_num,
            destination_file_name=destination_file_name,
            matrix_code=matrix_code, location=location,
            equipment_desc=equipment_desc

        )
        time.sleep(0.75)


async def get_files(main_path: str) -> list:
    """Получение списка файлов c расширением endswith из main_path

    :type main_path: str
    :param main_path: директория для поиска файлов
    """
    json_files_list = []
    for subdir, dirs, files in os.walk(main_path):
        for file in files:
            filepath = subdir + os.sep + file

            if filepath.endswith('.py'): continue
            if filepath.endswith('.jpg'): continue
            if filepath.endswith('.tmp'): continue
            if filepath.endswith('.db'): continue

            if not file: continue
            if '~' in file: continue

            matrix_code: str = await get_industrial_equipment_code(file)
            if not matrix_code: continue

            main_location: str = \
                subdir.replace(str(Udocan_media_path), '').replace(f'{os.sep}{SUB_LOCATION}', '').split(os.sep)[1]
            subdivision: str = \
                subdir.replace(str(Udocan_media_path), '').replace(f'{os.sep}{SUB_LOCATION}', '').split(os.sep)[2]

            matrix_description: str = ''

            json_files_list.append(
                {
                    "file": file,
                    "matrix_code": matrix_code,
                    "matrix_description": matrix_description,
                    "matrix_post": 'следующая установка',
                    "matrix_prev": 'предыдущая установка установка',
                    "main_location": main_location,
                    "glob_location": subdivision,
                    "location": 'литер здания',
                    "sub_location": 'участок или площадка',
                    "description": 'краткое описание установки',
                    "technology": 'краткое технологии',
                    "responsible": 'ответственный',
                    "subdir": subdir,
                    "full_file_path": str(Path(subdir, file)),
                }
            )
    return json_files_list


async def get_industrial_equipment_code(file_name: str) -> str:
    """

    :param file_name:
    :return:
    """
    name_list: list = file_name.split(' ')
    if len(name_list) == 0: return ''

    for item in name_list:
        part_post = ''
        part_pref = ''

        item = item.split('.')[0]

        if '-' not in item: continue

        if len(item.split('-')) < 2:
            logger.error(f'part_post: {item} {file_name}')

        try:
            if int(item.split('-')[-1]) > 0:
                part_post = item.split('-')[2]

        except (ValueError, IndexError) as err:
            logger.error(f'part_post: {item} {file_name}  {repr(err)}')
            continue

        try:
            if int(item.split('-')[0][2:4]):
                part_pref = item.split('-')[0]

        except ValueError as err:
            logger.error(f'part_pref: {item.split("-")[0]} {item} {file_name}  {repr(err)}')
            continue

        except IndexError as err:
            logger.error(f'part_post: {item} {file_name}  {repr(err)}')
            continue

        if all([part_pref, part_post]):
            return item

    # logger.info(f'{file_name = }')
    return ''


async def fanc_name() -> str:
    """Возвращает имя вызываемой функции"""
    stack = traceback.extract_stack()
    return str(stack[-2][2])


async def get_industrial_equipment_file_list(main_path: str):
    """"""
    if not os.path.isdir(main_path):
        os.makedirs(main_path)

    industrial_equipment_file_list: list = await get_files(main_path)
    logger.info(f'file_list: {len(industrial_equipment_file_list)}')
    return industrial_equipment_file_list


async def create_industrial_equipment_xlsx(chat_id, industrial_equipment_file_list, main_path):
    """

    :param chat_id:
    :param industrial_equipment_file_list:
    :param main_path:
    :return:
    """
    workbook, worksheet = await create_xlsx(chat_id, f'{main_path}isolation_matrix_report.xlsx')
    if not worksheet:
        return False

    for row, item in enumerate(industrial_equipment_file_list, start=1):
        try:
            worksheet.cell(row=row, column=1, value=row)
            worksheet.cell(row=row, column=2, value=item.get("file"))
            worksheet.cell(row=row, column=3, value=item.get("matrix_code"))
            worksheet.cell(row=row, column=4, value=item.get("matrix_post"))
            worksheet.cell(row=row, column=5, value=item.get("matrix_prev"))
            worksheet.cell(row=row, column=6, value=item.get("main_location"))
            worksheet.cell(row=row, column=7, value=item.get("subdivision"))
            worksheet.cell(row=row, column=8, value=item.get("location"))
            worksheet.cell(row=row, column=9, value=item.get("sub_location"))
            worksheet.cell(row=row, column=10, value=item.get("description"))
            worksheet.cell(row=row, column=11, value=item.get("technology"))
            worksheet.cell(row=row, column=12, value=item.get("responsible"))
            worksheet.cell(row=row, column=13, value=item.get("subdir"))

        except Exception as err:
            logger.error(f"set_user_values {repr(err)}")
            continue

    workbook.save(f'{main_path}isolation_matrix_report.xlsx')


async def create_main_folder(catalog_path, location):
    """

    :param catalog_path:
    :param location:
    :return:
    """

    if not os.path.isdir(str(Path(catalog_path, f'!{location}'))):
        os.makedirs(str(Path(catalog_path, f'!{location}')))

    catalog_location_path: list = [
        '00 Описание участка',
        '01 СОП',
        '02 СРП',
        '03 Техкарты',
        '88 Прочее',
        "99 Общее",
    ]

    for location_item in catalog_location_path:
        if not os.path.isdir(str(Path(catalog_path, f'!{location}', location_item))):
            os.makedirs(str(Path(catalog_path, f'!{location}', location_item)))


async def create_sub_folder(catalog_path, matrix_dir):
    """

    :return:
    """
    try:
        if not os.path.isdir(str(Path(catalog_path, matrix_dir))):
            os.makedirs(str(Path(catalog_path, matrix_dir)))
    except OSError as err:
        logger.error(f'create_sub_folder Синтаксическая ошибка в имени файла {repr(err)}')

    standard_folders_list: list = [
        "00 Описание и фото",
        "01 Технология",
        "02 База знаний",
        "03 Матрица изоляции",
        "04 Чертежи",
        "05 Схемы",
        "06 Чек-листы",
        "88 Прочее",
        "99 Общее по отделению",
    ]

    for folder_item in standard_folders_list:
        try:
            if not os.path.isdir(str(Path(catalog_path, matrix_dir, folder_item))):
                os.makedirs(str(Path(catalog_path, matrix_dir, folder_item)))
        except FileNotFoundError as err:
            logger.error(f'create_sub_folder FileNotFoundError {repr(err)}')


async def save_doc_file(src_file, destination_file: str):
    try:
        shutil.copyfile(src=src_file, dst=destination_file)
    except FileNotFoundError as err:
        logger.error(f'shutil.copyfile FileNotFoundError {repr(err)}')
        logger.error(f'{src_file = }')
        logger.error(f'{destination_file = }')


async def save_qr_code(qr_data: str, save_dir: str) -> str:
    """

    :return:
    """
    prefix = 'qr_matrix_nom_'
    if qr_data is None: return ''
    if qr_data == '': return ''

    words = f'{prefix}{qr_data}'
    save_name = f'{prefix}{qr_data}.jpg'

    colorized = False

    try:
        await create_qr_code(
            words,
            version=3,
            level='H',
            picture=None,
            colorized=colorized,
            contrast=2.0,
            brightness=2.0,
            save_name=save_name,
            save_dir=save_dir
        )
        return save_name

    except Exception as err:
        logger.error(f'{qr_data = } save_qr_code {repr(err)}')

        # for i in qr_data:
        #     try:
        #         lang = detect(i)
        #         print(f'{lang = } { i = }')
        #     except LangDetectException as err:
        #         print(f'{i} {err}')
        return ''


async def create_qr_xls_file(chat_id: str, qr_path: str, save_name: str, matrix_num: str,
                             destination_file_name: str, matrix_code: str, location: str,
                             equipment_desc: str) -> bool:
    """"""
    qr_doc_name: str = f'Тег_код {matrix_num} {destination_file_name}.xlsx'
    qr_doc_path: str = str(Path(qr_path, qr_doc_name))

    workbook, worksheet = await create_xlsx(chat_id, qr_doc_path)
    if not worksheet:
        await bot_send_message(chat_id=chat_id, text=f'{Messages.Error.worksheet_not_found} {matrix_code}')
        return False

    await set_act_page_setup(worksheet)

    row = 2
    row_range = [f'A{row}:K{row}']
    worksheet.merge_cells(f'A{row}:K{row}')

    for item, cell_range in enumerate(row_range, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    photographic_report_font = [
        [f'A{row}:K{row}', {"font_size": 32, "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    row_dimensions = [
        [f'{row}', 120],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    worksheet.cell(row=row, column=1, value=f'{matrix_num} {location}\n {matrix_code} \n{equipment_desc}')

    img_params: dict = {
        "photo_full_name": qr_doc_name,
        "height": 700,
        "width": 700,
        "anchor": True,
        "column": 'A',
        "column_img": 1,
        "row": 3,
    }
    service_image_name: str = f'{qr_path}{os.sep}{save_name}'
    await insert_service_image(worksheet, chat_id=chat_id, service_image_name=service_image_name, img_params=img_params)
    workbook.save(qr_doc_path)

    row = 40
    row_range = [f'B{row}:J{row}']
    worksheet.merge_cells(f'B{row}:J{row}')

    row_dimensions = [
        [f'{row}', 40],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    photographic_report_font = [
        [f'B{row}:J{row}', {"font_size": 30, "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    for item, cell_range in enumerate(row_range, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')
    worksheet.cell(row=row, column=2, value='Диспетчер ТК')

    row = 41
    row_range = [f'B{row}:J{row}']
    worksheet.merge_cells(f'B{row}:J{row}')

    row_dimensions = [
        [f'{row}', 40],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    photographic_report_font = [
        [f'B{row}:J{row}', {"font_size": 30, "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    for item, cell_range in enumerate(row_range, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')
    worksheet.cell(row=row, column=2, value='8-924-515-23-48')

    row = 42
    # row_range = [f'B{row}:J{row}']
    # worksheet.merge_cells(f'B{row}:J{row}')

    row_dimensions = [
        [f'{row}', 30],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    photographic_report_font = [
        [f'B{row}:J{row}', {"font_size": 14, "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    for item, cell_range in enumerate(row_range, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')
    worksheet.cell(row=row, column=1, value='доступ через TC_Assistant_Bot')

    worksheet.cell(row=row, column=11, value=f'{VERSION}')

    print_area = f'$A$1:K{42}'
    await set_act_page_after_footer_setup(worksheet, print_area)
    workbook.save(qr_doc_path)

    return True


async def set_row_dimensions(worksheet: Worksheet, row_number: int or str, height: int):
    """Установление высоты строки

    :param worksheet:
    :param row_number:
    :param height:
    :return:
    """
    try:
        worksheet.row_dimensions[int(row_number)].height = float(height)
    except Exception as err:
        logger.error(f"set_row_dimensions {repr(err)}")


async def sets_report_font(worksheet, cell_range, params: dict) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(
                color=params.get("color"),
                italic=params.get("italic"),
                size=params.get("font_size"),
                bold=params.get("bold"),
                name=params.get("name"),
                underline=params.get("underline"),
                vertAlign=params.get("vertAlign"),
            )

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"sets_report_font {repr(err)}")
            continue
    return True


async def set_act_alignment(worksheet, cell_range, horizontal=None, vertical=None):
    """Форматирование ячейки: положение текста в ячейке (лево верх)

    """
    wrap_alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.alignment = wrap_alignment

        except Exception as err:
            logger.error(f"iter {item} cell {cell}")
            logger.error(f"set_mip_alignment {repr(err)}")


async def set_act_page_setup(worksheet: Worksheet) -> bool:
    """Установка параметров страницы

    :param worksheet:
    :return: bool
    """
    #  https://xlsxwriter.readthedocs.io/page_setup.html
    #  worksheet.print_title_rows = '$2:$3'
    #  worksheet.print_title = '$2:$3'

    # Printer Settings
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

    # Подогнать область печати к определенному кол-у страниц как по вертикали, так и по горизонтали.
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = '0'

    # worksheet.views
    worksheet.print_options.horizontalCentered = True
    worksheet.print_area = '$A$1:K50'

    #  масштабный коэффициент для распечатываемой страницы
    #  worksheet.set_print_scale(75)

    #  worksheet.row_breaks.append(Break(id=53))
    #  worksheet.col_breaks.append(Break(id=13))

    # задаем собственные значения отступов
    top_bottom = 2 / 2.54
    left_right = 2 / 2.54
    worksheet.page_margins = PageMargins(left=left_right, right=left_right, top=top_bottom, bottom=top_bottom)

    # worksheet.oddFooter.left.text = "Страница &[Page] из &N"
    # worksheet.oddFooter.left.size = 10
    # worksheet.oddFooter.left.font = "Arial,Bold"
    # worksheet.oddFooter.left.color = "030303"
    worksheet.differentFirst = False
    worksheet.differentOddEven = True

    return True


async def set_act_page_after_footer_setup(worksheet: Worksheet, print_area: str, break_line: int = None) -> bool:
    """Установка параметров страницы

    :param break_line: int
    :param print_area: str
    :param worksheet: Worksheet
    :return: bool
    """

    #  https://xlsxons.horizontalCentered = True
    worksheet.print_area = print_area

    #  масштабный коэффициент для распечатываемой страницы
    # worksheet.set_print_scale(75)

    # задаем собственные значения отступов
    top_bottom = 2 / 2.54
    left_right = 2 / 2.54
    # worksheet.page_margins = PageMargins(left=left_right, right=left_right, top=top_bottom, bottom=top_bottom)
    worksheet.page_margins = PageMargins()

    if break_line:
        worksheet.row_breaks.append(Break(id=break_line))
        # worksheet.col_breaks.append(Break(id=13))

    return True


async def insert_service_image(worksheet: Worksheet, *, chat_id: str = None, service_image_name: str = None,
                               img_params: dict = None) -> bool:
    """Вставка изображений в файл

    :param service_image_name: str - имя файла для обработки
    :param chat_id: int - id пользователя (папки) где находится logo
    :param img_params: dict параметры вставки
    :param worksheet:
    :return: bool
    """

    if not os.path.isfile(service_image_name):
        logger.error("service image not found")
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_create)
        return False

    img: Image = Image(service_image_name)

    img = await image_preparation(img, img_params)

    result = await insert_images(worksheet, img=img)
    if result:
        return True

    return False


async def image_preparation(img: Image, img_params: dict):
    """Подготовка изображения перед вставкой на страницу
    """

    height = img_params.get("height")
    scale = img_params.get("scale")
    width = img_params.get("width")
    anchor = img_params.get("anchor")
    column = img_params.get("column")
    row = img_params.get("row")

    # высота изображения
    if height:
        img.height = height

    # ширина изображения
    if width:
        img.width = width

    # изменение пропорций изображения
    if scale:
        scale = img.height / max(img.height, img.width)
        img.width = img.width * scale

    # прикрепление изображение по адресу str(column) + str(row)
    if anchor:
        img.anchor = str(column) + str(row)

    return img


async def insert_images(worksheet: Worksheet, img: Image) -> bool:
    """Вставка изображения на лист worksheet*,
    :param img: файл изображения
    :param worksheet: Worksheet - страница документа для вставки изображения
    :return:
    """

    try:
        worksheet.add_image(img)
        return True

    except Exception as err:
        logger.error(f"insert_images {repr(err)}")
        return False


async def check_dataframe(dataframe: DataFrame, hse_user_id: str | int) -> bool:
    """Проверка dataframe на наличие данных

    :param dataframe:
    :param hse_user_id: id пользователя
    :return:
    """
    if dataframe is None:
        text_violations: str = 'не удалось получить данные!'
        logger.error(f'{hse_user_id = } {text_violations}')
        return False

    if dataframe.empty:
        return False

    return True


async def create_xlsx(chat_id: str, full_act_path: str):
    """
    """

    is_created: bool = await create_new_xlsx(report_file=full_act_path)
    if is_created is None:
        logger.warning(Messages.Error.workbook_not_create)
        # await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_create)
        return None, None

    workbook = await get_workbook(fill_report_path=full_act_path)
    if workbook is None:
        logger.warning(Messages.Error.workbook_not_found)
        # await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_found)
        return None, None

    worksheet = await get_worksheet(workbook, index=0)
    if worksheet is None:
        logger.warning(Messages.Error.worksheet_not_found)
        # await bot_send_message(chat_id=chat_id, text=Messages.Error.worksheet_not_found)
        return None, None

    return workbook, worksheet


async def create_new_xlsx(report_file: str) -> bool:
    """Создание xlsx
    """
    try:
        wb = openpyxl.Workbook()
        wb.save(report_file)
        return True

    except (TypeError, Exception) as err:
        logger.error(F"set_border {repr(err)}")
        return False


async def get_worksheet(wb: Workbook, index: int = 0) -> Worksheet:
    """Получение Страницы из документа по индексу
    :param wb: Workbook - книга xls
    :param index: int - индекс листа
    :return: worksheet or None
    """
    try:
        worksheet: Worksheet = wb.worksheets[index]
        return worksheet

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def get_workbook(fill_report_path: str) -> Workbook | None:
    """Открыть и загрузить Workbook
    :param fill_report_path: str полный путь к файлу
    :return: Workbook or None
    """
    try:
        workbook: Workbook = openpyxl.load_workbook(fill_report_path)
        return workbook

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def test_2():
    chat_id: str = str(373084462)

    save_dir: str = 'C:\\Users\\KDeusEx\\PycharmProjects\\!media\\!industrial_equipment_catalog\\0001 ZF01-FT-101 УИиО ФлотомашинаTANKCELL E200\\88 Прочее\\'
    destination_file_name: str = 'ZF01-FT-101 УИиО ФлотомашинаTANKCELL E200'
    save_name: str = 'qr_matrix_nom_ZF01-FT-101.jpg'
    matrix_num: str = '0001'
    location: str = 'УИиО'

    qr_path: str = ''
    matrix_code: str = ''
    equipment_desc: str = ''

    await create_qr_xls_file(
        chat_id=chat_id,
        qr_path=save_dir,
        save_name=save_name,
        matrix_num=matrix_num,
        destination_file_name=destination_file_name,
        location=location,

    )

    return ''


async def test():
    chat_id: str = str(373084462)
    main_path: str = str(Path(Udocan_media_path, SUB_LOCATION))

    industrial_equipment_file_list: list = await get_industrial_equipment_file_list(main_path)

    # await create_industrial_equipment_xlsx(chat_id, industrial_equipment_file_list, main_path)

    await create_industrial_equipment_catalog(chat_id, industrial_equipment_file_list, str(Udocan_media_path))

    return ''


if __name__ == '__main__':
    asyncio.run(test())
